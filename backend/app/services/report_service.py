import csv
import io
from typing import List, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

class ReportService:
    def generate_csv_report(self, crimes: List[Dict[str, Any]]) -> str:
        """
        Generate CSV string of crime records.
        """
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        writer.writerow([
            "Crime ID", "FIR Number", "Crime Type", "Category", 
            "Subcategory", "Date", "Time", "District", 
            "Police Station", "Latitude", "Longitude", 
            "Victim Name", "Victim Age", "Victim Gender",
            "Suspect Name", "Status", "IPC/BNS Sections", "Description"
        ])
        
        for crime in crimes:
            date_val = crime.get("date")
            if isinstance(date_val, datetime):
                date_val = date_val.strftime("%Y-%m-%d")
                
            victim = crime.get("victim", {})
            suspect = crime.get("suspect", {})
            sections = ", ".join(crime.get("sections", []))
            
            writer.writerow([
                crime.get("crime_id"),
                crime.get("FIR_number"),
                crime.get("crime_type"),
                crime.get("crime_category"),
                crime.get("crime_subcategory"),
                date_val,
                crime.get("time"),
                crime.get("district"),
                crime.get("police_station"),
                crime.get("latitude"),
                crime.get("longitude"),
                victim.get("name", ""),
                victim.get("age", ""),
                victim.get("gender", ""),
                suspect.get("name", ""),
                crime.get("status"),
                sections,
                crime.get("description")
            ])
            
        return output.getvalue()

    def generate_excel_report(self, crimes: List[Dict[str, Any]]) -> bytes:
        """
        Generate Excel bytes of crime records.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Crime Records Summary"
        
        # Color palettes
        navy_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        title_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        
        headers = [
            "Crime ID", "FIR Number", "Crime Type", "Category", 
            "Date", "District", "Police Station", "Suspect", "Status"
        ]
        
        ws.append(headers)
        
        # Format headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = title_font
            cell.fill = navy_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Append data
        for r_idx, crime in enumerate(crimes, start=2):
            date_val = crime.get("date")
            if isinstance(date_val, datetime):
                date_val = date_val.strftime("%Y-%m-%d")
                
            suspect = crime.get("suspect", {})
            
            ws.append([
                crime.get("crime_id"),
                crime.get("FIR_number"),
                crime.get("crime_type"),
                crime.get("crime_category"),
                date_val,
                crime.get("district"),
                crime.get("police_station"),
                suspect.get("name", "Unknown"),
                crime.get("status")
            ])
            
        # Autofit column widths
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_to_check = str(cell.value or '')
                if len(val_to_check) > max_len:
                    max_len = len(val_to_check)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def generate_pdf_report(self, crimes: List[Dict[str, Any]], title: str = "Crime Database Report") -> bytes:
        """
        Generate ReportLab PDF bytes of crime records.
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=40,
            bottomMargin=40
        )
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            name="ReportTitle",
            parent=styles["Title"],
            fontName="Helvetica-Bold",
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#1F4E79"),
            alignment=0, # Left-aligned
            spaceAfter=15
        )
        
        meta_style = ParagraphStyle(
            name="ReportMeta",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#555555"),
            spaceAfter=20
        )
        
        cell_style = ParagraphStyle(
            name="TableCell",
            parent=styles["Normal"],
            fontName="Helvetica",
            fontSize=8,
            leading=10
        )
        
        header_cell_style = ParagraphStyle(
            name="TableHeaderCell",
            parent=styles["Normal"],
            fontName="Helvetica-Bold",
            fontSize=8,
            leading=10,
            textColor=colors.white
        )

        elements = []
        
        # Add Header
        elements.append(Paragraph(title, title_style))
        elements.append(Paragraph(f"Generated on: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}<br/>Total Records: {len(crimes)}", meta_style))
        elements.append(Spacer(1, 10))
        
        # Table of Crime Records
        # Headers: FIR, Date, Type, District, Status
        table_data = [[
            Paragraph("FIR Number", header_cell_style),
            Paragraph("Date", header_cell_style),
            Paragraph("Crime Type", header_cell_style),
            Paragraph("District", header_cell_style),
            Paragraph("Police Station", header_cell_style),
            Paragraph("Status", header_cell_style)
        ]]
        
        for c in crimes[:100]:  # Limit to first 100 to prevent massive PDFs during export
            date_val = c.get("date")
            if isinstance(date_val, datetime):
                date_val = date_val.strftime("%Y-%m-%d")
            else:
                date_val = str(date_val)[:10]
                
            table_data.append([
                Paragraph(c.get("FIR_number", ""), cell_style),
                Paragraph(date_val, cell_style),
                Paragraph(c.get("crime_type", ""), cell_style),
                Paragraph(c.get("district", ""), cell_style),
                Paragraph(c.get("police_station", ""), cell_style),
                Paragraph(c.get("status", ""), cell_style)
            ])
            
        # Table styling
        # Column widths: total width = 8.5 * 72 - 72 = 540
        col_widths = [90, 70, 100, 90, 110, 80]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F4E79")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('TOPPADDING', (0,0), (-1,0), 6),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.HexColor("#F9FBFD"), colors.white]),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#D0D8E0")),
            ('TOPPADDING', (0,1), (-1,-1), 5),
            ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ]))
        
        elements.append(t)
        
        # Build Document
        doc.build(elements)
        return buffer.getvalue()
